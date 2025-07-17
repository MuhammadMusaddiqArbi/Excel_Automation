import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

class ExcelAutomation:
    wb = None
    def create_workbook(filename):
        ExcelAutomation.wb = Workbook()
        print(f"Workbook created {filename}.")

    def save_workbook(filename):
        if ExcelAutomation.wb is not None:
            ExcelAutomation.wb.save(filename)
            print(f"Workbook saved as {filename}.")
        else:
            print("No workbook to save. Please create a workbook first.")

    def load_workbook(filename):
        try:
            ExcelAutomation.wb = load_workbook(filename)
            print(f"Workbook {filename} loaded successfully.")
        except FileNotFoundError:
            print(f"Workbook {filename} not found.")
        except Exception as e:
            print(f"An error occurred while loading the workbook: {e}")

    def add_sheet(sheet_name):
        if ExcelAutomation.wb is not None:
            ExcelAutomation.wb.create_sheet(title=sheet_name)
            print(f"Sheet '{sheet_name}' added to the workbook.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def delete_sheet(sheet_name):
        if ExcelAutomation.wb is not None:
            ExcelAutomation.wb.remove(ExcelAutomation.wb[sheet_name])
            print(f"Sheet '{sheet_name}' deleted from the workbook.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def list_sheets():
        if ExcelAutomation.wb is not None:
            sheets = ExcelAutomation.wb.sheetnames
            print(sheets)
        else:
            print("No workbook available. Please create or load a workbook first.")

    def write_data(sheet_name, cell, data):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            sheet[cell] = data 
            print("Data written to sheet successfully.")
        else: 
            print("No workbook available. Please create or load a workbook first.")

    def read_data(sheet_name, cell):
        if ExcelAutomation.wb is not None:
            sheet = ExcelAutomation.wb[sheet_name]
            data = sheet[cell].value
            print(f"Data: {data}")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def close_workbook():
        if ExcelAutomation.wb is not None:
            ExcelAutomation.wb.close()
            print("Workbook closed.")
        else:
            print("No workbook to close. Please create or load a workbook first.")    

    def create_chart(sheet_name, chart_type, data_range, title):
        if ExcelAutomation.wb is not None:
            sheet = ExcelAutomation.wb[sheet_name]
            if chart_type == 'bar':
                chart = BarChart()
            elif chart_type == 'line':
                chart = LineChart()
            elif chart_type == 'pie':
                chart = PieChart()
            else:
                print("Unsupported chart type.")
                return
            
            data = Reference(sheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            chart.title = title
            sheet.add_chart(chart, "E5")

    def save_chart(sheet_name, chart_type, data_range, title, filename):
        ExcelAutomation.create_chart(sheet_name, chart_type, data_range, title)
        ExcelAutomation.save_workbook(filename)
        print(f"Chart saved in {filename}.")

    def read_chart(sheet_name, chart_type):
        if ExcelAutomation.wb is not None:
            sheet = ExcelAutomation.wb[sheet_name]
            charts = sheet._charts
            for chart in charts:
                if chart.type == chart_type:
                    print(f"Chart of type {chart_type} found.")
                    return chart
            print(f"No chart of type {chart_type} found in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def delete_chart(sheet_name, chart_type):
        if ExcelAutomation.wb is not None:
            sheet = ExcelAutomation.wb[sheet_name]
            charts = sheet._charts
            for chart in charts:
                if chart.type == chart_type:
                    sheet._charts.remove(chart)
                    print(f"Chart of type {chart_type} deleted from sheet {sheet_name}.")
                    return
            print(f"No chart of type {chart_type} found in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def merge_cells(sheet_name, start_cell, end_cell):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            sheet.merge_cells(f"{start_cell}:{end_cell}")
            print(f"Cells merged from {start_cell} to {end_cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def unmerge_cells(sheet_name, start_cell, end_cell):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            sheet.unmerge_cells(f"{start_cell}:{end_cell}")
            print(f"Cells unmerged from {start_cell} to {end_cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def freeze_panes(sheet_name, cell):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            sheet.freeze_panes = cell
            print(f"Frozen panes set at {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def unfreeze_panes(sheet_name):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            sheet.freeze_panes = None
            print(f"Unfrozen panes in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_column_width(sheet_name, column, width):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            sheet.column_dimensions[column].width = width
            print(f"Set width of column {column} to {width} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_row_height(sheet_name, row, height):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            sheet.row_dimensions[row].height = height
            print(f"Set height of row {row} to {height} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def format_cell(sheet_name, cell, font=None, fill=None, border=None, alignment=None):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            if font:
                cell_obj.font = font
            if fill:
                cell_obj.fill = fill
            if border:
                cell_obj.border = border
            if alignment:
                cell_obj.alignment = alignment
            print(f"Formatted cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_cell_style(sheet_name, cell, style):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.style = style
            print(f"Set style of cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_cell_comment(sheet_name, cell, comment):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.comment = comment
            print(f"Set comment for cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.") 

    def set_cell_validation(sheet_name, cell, validation):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.data_validation = validation
            print(f"Set data validation for cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_cell_hyperlink(sheet_name, cell, url):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.hyperlink = url
            print(f"Set hyperlink for cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_cell_protection(sheet_name, cell, locked=True):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.protection = {'locked': locked}
            print(f"Set protection for cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_cell_number_format(sheet_name, cell, number_format):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.number_format = number_format
            print(f"Set number format for cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_cell_formula(sheet_name, cell, formula):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.value = formula
            print(f"Set formula for cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")

    def set_cell_style(sheet_name, cell, style):
        if ExcelAutomation.wb is not None:
            if sheet_name not in ExcelAutomation.wb.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist. Please create it first.")
                return
            sheet = ExcelAutomation.wb[sheet_name]
            cell_obj = sheet[cell]
            cell_obj.style = style
            print(f"Set style for cell {cell} in sheet {sheet_name}.")
        else:
            print("No workbook available. Please create or load a workbook first.")
                                                                                                         

    def continuee():
        print("Do You want to continue? (Y/N)")
        choice = input().strip().upper()
        if choice == 'Y':
            main()
        elif choice == 'N':
            print("Exiting the script.")
            return 0
        else:
            print("Invalid choice. Please enter 'Y' or 'N'.")
            ExcelAutomation.continuee()    



def main():
    print("Excel Automation Script")
    print("1. Create Workbook")
    print("2. Save Workbook")
    print("3. Load Workbook")
    print("4. Add Sheet")
    print("5. Delete Sheet")
    print("6. List Sheets")
    print("7. Write Data")
    print("8. Read Data")
    print("9. Close Workbook")
    print("10. Create Chart")
    print("11. Save Chart")
    print("12. Read Chart")
    print("13. Delete Chart")
    print("14. Merge Cells")
    print("15. Unmerge Cells")
    print("16. Freeze Panes")
    print("17. Unfreeze Panes")
    print("18. Set Column Width")
    print("19. Set Row Height")
    print("20. Format Cell")
    print("21. Set Cell Style")
    print("22. Set Cell Comment")
    print("23. Set Cell Validation")
    print("24. Set Cell Hyperlink")
    print("25. Set Cell Protection")
    print("26. Set Cell Number Format")
    print("27. Set Cell Formula")
    print("28. Set Cell Style")
    print("29. Exit")
    choice = input("Enter your choice: ")
    while choice != '29':
        if choice == '1':
            filename = input("Enter filename to create: ")
            ExcelAutomation.create_workbook(filename)
            ExcelAutomation.continuee()
        elif choice == '2':
            filename = input("Enter filename to save: ")
            ExcelAutomation.save_workbook(filename)
            ExcelAutomation.continuee()
        elif choice == '3':
            filename = input("Enter filename to load: ")
            ExcelAutomation.load_workbook(filename)
            ExcelAutomation.continuee()
        elif choice == '4':
            sheet_name = input("Enter sheet name to add: ")
            ExcelAutomation.add_sheet(sheet_name)
            ExcelAutomation.continuee() 
        elif choice == '5':
            sheet_name = input("Enter sheet name to delete: ")
            ExcelAutomation.delete_sheet(sheet_name)
            ExcelAutomation.continuee()
        elif choice == '6':
            ExcelAutomation.list_sheets()
            ExcelAutomation.continuee()
        elif choice == '7':
            sheet_name = input("Enter sheet name: ")
            cell = input("Enter cell (e.g., A1): ")
            data = input("Enter data to write: ")
            ExcelAutomation.write_data(sheet_name, cell, data)
            ExcelAutomation.continuee()
        elif choice == '8':
            sheet_name = input("Enter sheet name: ")
            cell = input("Enter cell (e.g., A1): ")
            ExcelAutomation.read_data(sheet_name, cell)
            ExcelAutomation.continuee()
        elif choice == '9':
            ExcelAutomation.close_workbook()
            ExcelAutomation.continuee()
        elif choice == '10':
            sheet_name = input("Enter sheet name for chart: ")
            chart_type = input("Enter chart type (bar/line/pie): ")
            data_range = input("Enter data range (e.g., A1:B10): ")
            title = input("Enter chart title: ")
            ExcelAutomation.create_chart(sheet_name, chart_type, data_range, title)
            print("Chart created successfully.")
            ExcelAutomation.continuee()
        elif choice == '11':
            sheet_name = input("Enter sheet name for chart: ")
            chart_type = input("Enter chart type (bar/line/pie): ")
            data_range = input("Enter data range (e.g., A1:B10): ")
            title = input("Enter chart title: ")
            filename = input("Enter filename to save the workbook with chart: ")
            ExcelAutomation.save_chart(sheet_name, chart_type, data_range, title, filename)
            print("Chart saved successfully.")
            ExcelAutomation.continuee()
        elif choice == '12':
            sheet_name = input("Enter sheet name to read chart from: ")
            chart_type = input("Enter chart type to read (bar/line/pie): ")
            ExcelAutomation.read_chart(sheet_name, chart_type)
            ExcelAutomation.continuee()
        elif choice == '13':
            sheet_name = input("Enter sheet name to delete chart from: ")
            chart_type = input("Enter chart type to delete (bar/line/pie): ")
            ExcelAutomation.delete_chart(sheet_name, chart_type)
            ExcelAutomation.continuee()
        elif choice == '14':
            sheet_name = input("Enter sheet name to merge cells: ")
            start_cell = input("Enter start cell (e.g., A1): ")
            end_cell = input("Enter end cell (e.g., B2): ")
            ExcelAutomation.merge_cells(sheet_name, start_cell, end_cell)
            ExcelAutomation.continuee()
        elif choice == '15':
            sheet_name = input("Enter sheet name to unmerge cells: ")
            start_cell = input("Enter start cell (e.g., A1): ")
            end_cell = input("Enter end cell (e.g., B2): ")
            ExcelAutomation.unmerge_cells(sheet_name, start_cell, end_cell)
            ExcelAutomation.continuee()
        elif choice == '16':
            sheet_name = input("Enter sheet name to freeze panes: ")
            cell = input("Enter cell to freeze panes at (e.g., A2): ")
            ExcelAutomation.freeze_panes(sheet_name, cell)
            ExcelAutomation.continuee()
        elif choice == '17':
            sheet_name = input("Enter sheet name to unfreeze panes: ")
            ExcelAutomation.unfreeze_panes(sheet_name)
            ExcelAutomation.continuee()
        elif choice == '18':
            sheet_name = input("Enter sheet name to set column width: ")
            column = input("Enter column (e.g., A): ")
            width = float(input("Enter width: "))
            ExcelAutomation.set_column_width(sheet_name, column, width)
            ExcelAutomation.continuee()
        elif choice == '19':
            sheet_name = input("Enter sheet name to set row height: ")
            row = int(input("Enter row number (e.g., 1): "))
            height = float(input("Enter height: "))
            ExcelAutomation.set_row_height(sheet_name, row, height)
            ExcelAutomation.continuee()
        elif choice == '20':
            sheet_name = input("Enter sheet name to format cell: ")
            cell = input("Enter cell (e.g., A1): ")
            font = input("Enter font (e.g., Arial, 12, bold): ")
            fill = input("Enter fill color (e.g., yellow): ")
            border = input("Enter border style (e.g., thin): ")
            alignment = input("Enter alignment (e.g., center): ")
            ExcelAutomation.format_cell(sheet_name, cell, font, fill, border, alignment)
            ExcelAutomation.continuee()
        elif choice == '21':
            sheet_name = input("Enter sheet name to set cell style: ")
            cell = input("Enter cell (e.g., A1): ")
            style = input("Enter style (e.g., currency): ")
            ExcelAutomation.set_cell_style(sheet_name, cell, style)
            ExcelAutomation.continuee()
        elif choice == '22':    
            sheet_name = input("Enter sheet name to set cell comment: ")
            cell = input("Enter cell (e.g., A1): ")
            comment = input("Enter comment text: ")
            ExcelAutomation.set_cell_comment(sheet_name, cell, comment)
            ExcelAutomation.continuee()  
        elif choice == '23':
            sheet_name = input("Enter sheet name to set cell validation: ")
            cell = input("Enter cell (e.g., A1): ")
            validation = input("Enter validation rule (e.g., list, 1,2,3): ")
            ExcelAutomation.set_cell_validation(sheet_name, cell, validation)
            ExcelAutomation.continuee()
        elif choice == '24':
            sheet_name = input("Enter sheet name to set cell hyperlink: ")
            cell = input("Enter cell (e.g., A1): ")
            url = input("Enter URL: ")
            ExcelAutomation.set_cell_hyperlink(sheet_name, cell, url)
            ExcelAutomation.continuee()
        elif choice == '25':
            sheet_name = input("Enter sheet name to set cell protection: ")
            cell = input("Enter cell (e.g., A1): ")
            locked = input("Lock cell? (yes/no): ").strip().lower() == 'yes'
            ExcelAutomation.set_cell_protection(sheet_name, cell, locked)
            ExcelAutomation.continuee()
        elif choice == '26':
            sheet_name = input("Enter sheet name to set cell number format: ")
            cell = input("Enter cell (e.g., A1): ")
            number_format = input("Enter number format (e.g., #,##0.00): ")
            ExcelAutomation.set_cell_number_format(sheet_name, cell, number_format)
            ExcelAutomation.continuee()
        elif choice == '27':
            sheet_name = input("Enter sheet name to set cell formula: ")
            cell = input("Enter cell (e.g., A1): ")
            formula = input("Enter formula (e.g., =SUM(B1:B10)): ")
            ExcelAutomation.set_cell_formula(sheet_name, cell, formula)
            ExcelAutomation.continuee()
        elif choice == '28':
            sheet_name = input("Enter sheet name to set cell style: ")
            cell = input("Enter cell (e.g., A1): ")
            style = input("Enter style (e.g., currency): ")
            ExcelAutomation.set_cell_style(sheet_name, cell, style)
            ExcelAutomation.continuee()                                                      
        elif choice == '29':
            print("Exiting the script.")
            return 0
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()